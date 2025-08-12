"""**************************************************************

Biblioteca para tomada de decisão entre a câmera, RFID e potenciômetro
Autor: Alex Luiz Cibien E Silva
Data de início: setembro de 2024

**************************************************************"""

from datetime import datetime
from os import path
from time import sleep, time

from openpyxl import Workbook, load_workbook


class DecisionMkg:
    """Classe para tomada de decisão entre a câmera, RFID e potenciômetro"""

    pot_reading: int = 0
    pot_value: int = 0
    camera_value: int = 0
    rfid_value: int = 0
    vel_value: int = 0
    vel_real_value: int = 0
    car_had_stopped: bool = False
    last_detected_sign: int = 0
    previous_detected_sign: int = 0

    PEDESTRE = 5
    VEL20KMH = 4
    PARAR = 3
    OCIOSO = 2
    N_DETECTOU = 2
    N_LIMITE = 2
    ACELERAR = 1

    def __init__(self):
        print("iniciando biblioteca Decision Making")

    def get_cam_value(self, _cam_value) -> int:
        """Função para obter o valor da câmera"""
        if _cam_value == "20km/h":
            self.camera_value = self.VEL20KMH
        elif _cam_value == "PARE":
            self.camera_value = self.PARAR
        elif _cam_value == "":
            self.camera_value = self.N_DETECTOU

        print(f"O valor da camera é: {self.camera_value}")

        return self.camera_value

    def get_pot_value(self, _pot_reading) -> int:
        """Função para obter o valor do potenciômetro"""
        if _pot_reading >= 10:
            self.pot_value = self.ACELERAR
        elif _pot_reading <= -10:
            self.pot_value = self.PARAR
        else:
            self.pot_value = self.OCIOSO

        print(f"O valor do potenciometro é: {self.pot_value}")

        return self.pot_value

    def get_rfid_value(self) -> int:
        """Função para obter o valor do RFID"""
        if self.rfid_value == self.VEL20KMH:
            self.rfid_value = self.VEL20KMH
        elif self.rfid_value == self.N_DETECTOU:
            self.rfid_value = self.N_DETECTOU
        elif self.rfid_value == self.PEDESTRE:
            self.rfid_value = self.PEDESTRE
        elif self.rfid_value == self.PARAR:
            self.rfid_value = self.PARAR

        print(f"O valor do RFID é: {self.rfid_value}")

        return self.rfid_value

    def update_detected_sign(self, new_sign: int) -> None:
        """Função para atualizar o sinal detectado"""
        self.previous_detected_sign = self.last_detected_sign
        self.last_detected_sign = new_sign

    def brake_and_go_command(self, reading_from_pot: int) -> None:
        """Função para frear e aguardar 5 segundos antes de acelerar novamente"""
        if self.car_had_stopped:
            self.vel_value = 0
            sleep(5)  # Aguarda cinco segundos por segurança

            if reading_from_pot == self.ACELERAR:
                return
            else:
                self.car_had_stopped = False

    def decision_from_input_improved(
        self, value_from_pot: int, value_from_cam: int, value_from_rfid: int
    ) -> int:
        """Função para tomar decisões baseadas nos valores de entrada"""
        if self.PARAR in (value_from_pot, value_from_cam, value_from_rfid):
            print(
                f"POT {value_from_pot}, CAM {value_from_cam} e RFID {value_from_rfid}"
            )
            return self.PARAR
        if value_from_rfid == self.PEDESTRE:
            print(
                f"POT {value_from_pot}, CAM {value_from_cam} e RFID {value_from_rfid}"
            )
            return self.PEDESTRE

        if value_from_pot in (self.ACELERAR, self.OCIOSO):
            print(f"POT {value_from_pot}")
            if value_from_cam == self.N_DETECTOU:
                print(f"CAM {value_from_cam}")
                if value_from_rfid == self.N_DETECTOU:
                    print(f"RFID {value_from_rfid}")
                    return self.N_LIMITE
                if value_from_rfid == self.VEL20KMH:
                    print(f"RFID {value_from_rfid}")
                    return self.VEL20KMH

            if value_from_cam == self.VEL20KMH:
                print(f"CAM {value_from_cam}")
                if value_from_rfid == self.N_DETECTOU:
                    print(f"RFID {value_from_rfid}")
                    return self.N_LIMITE
                if value_from_rfid == self.VEL20KMH:
                    print(f"RFID {value_from_rfid}")
                    return self.VEL20KMH

        return self.PARAR

    # simulação de montecarlo - jogar dados aleatórios
    # 100000 dados
    # utilizar uma lista com as ações e respostas em ordem aleatória, isto é não precisa ser 1,2,3...22, o importante é ser o mais aleatório possível
    # a simulação de montecarlo pode ser utilizada para validar
    # a entrada é o "montecarlo" e a saída é o meu programa (whichIndexMatrix)
    # da de fazer tudo pelo pc

    def which_index_matrix(
        self, value_from_pot: int, value_from_cam: int, value_from_rfid: int
    ) -> int:
        """Função para determinar o índice da matriz de decisão"""
        if value_from_pot == self.PARAR:
            if value_from_cam == self.PARAR:
                if value_from_rfid == self.PARAR:
                    return 1
                if value_from_rfid == self.PEDESTRE:
                    return 2
                if value_from_rfid == self.VEL20KMH:
                    return 3
            elif value_from_cam == self.N_DETECTOU:
                if value_from_rfid == self.N_DETECTOU:
                    return 4
                if value_from_rfid == self.PARAR:
                    return 5
                if value_from_rfid == self.VEL20KMH:
                    return 6
                if value_from_rfid == self.PEDESTRE:
                    return 7
            elif value_from_cam == self.VEL20KMH:
                if value_from_rfid == self.PARAR:
                    return 8
                elif value_from_rfid == self.PEDESTRE:
                    return 9
                elif value_from_rfid == self.VEL20KMH:
                    return 10

        elif value_from_pot in (self.ACELERAR, self.OCIOSO):
            if value_from_cam == self.PARAR:
                if value_from_rfid == self.PARAR:
                    return 11
                if value_from_rfid == self.PEDESTRE:
                    return 12
                if value_from_rfid == self.VEL20KMH:
                    return 13
            elif value_from_cam == self.N_DETECTOU:
                if value_from_rfid == self.N_DETECTOU:
                    return 14
                if value_from_rfid == self.PARAR:
                    return 15
                if value_from_rfid == self.PEDESTRE:
                    return 16
                if value_from_rfid == self.VEL20KMH:
                    return 17
            elif value_from_cam == self.VEL20KMH:
                if value_from_rfid == self.VEL20KMH:
                    return 18
                if value_from_rfid == self.PEDESTRE:
                    return 19
                if value_from_rfid == self.PARAR:
                    return 20
        return 0

    def execute_decision(self, value_from_decision: int) -> None:
        """Função para executar a decisão tomada"""
        if value_from_decision == self.PARAR:
            self.brake_and_go_command(self.pot_reading)
        elif value_from_decision == self.N_LIMITE:
            self.vel_value == self.pot_reading / 2
        elif value_from_decision == self.VEL20KMH:
            if self.vel_real_value <= 18:
                print("deixa o comando do motorista")
                self.vel_value == self.pot_reading / 2
            elif self.vel_real_value >= 22:
                print("abaixa pra 20km/hr")
                self.vel_value == 20
        print(self.vel_value)

    def initialize_excel(self, filename="log_data.xlsx"):
        """Função para inicializar o arquivo Excel"""
        if path.exists(filename):
            excel_workbook = load_workbook(filename)

        else:
            excel_workbook = Workbook()
            sheet = excel_workbook.active
            if not sheet:
                return
            sheet.title = "DataLog"
            sheet.append(
                [
                    "Timestamp",
                    "Potenciometro",
                    "Camera",
                    "RFID",
                    "index de Entrada",
                    "index de Saida",
                ]
            )
            excel_workbook.save(filename)

        return excel_workbook

    def save_log(
        self,
        workbook_save,
        _timestamp: str,
        value_from_pot: int,
        value_from_cam: int,
        value_from_rfid: int,
        index_from_enter,
        index_from_exit,
        filename="log_data.xlsx",
    ) -> None:
        """Função para salvar os dados no arquivo Excel"""
        sheet = workbook_save.active
        sheet.append(
            [
                _timestamp,
                value_from_pot,
                value_from_cam,
                value_from_rfid,
                index_from_enter,
                index_from_exit,
            ]
        )
        workbook_save.save(filename)
        print(f"Dados salvos no Excel com timestamp {_timestamp}")


if __name__ == "__main__":
    from Sign_Detection_TCC.detection import capture_photo, getClassName, indexVal

    INTERVALO = 1
    ultimo_tempo = time()
    app = DecisionMkg()
    workbook = app.initialize_excel()

    while True:
        tempo_atual = time()
        raw_cam_value = getClassName(indexVal)
        potentiometer_reading = app.pot_reading
        potentiometer_value = app.get_pot_value(potentiometer_reading)
        cam_value = app.get_cam_value(raw_cam_value)
        rfid_value = app.get_rfid_value()

        if tempo_atual - ultimo_tempo >= INTERVALO:
            decision_from_logic = app.decision_from_input_improved(
                potentiometer_value, cam_value, rfid_value
            )
            # app.executeDecision(decisionFromLogic)
            indexExit = app.which_index_matrix(
                potentiometer_value, cam_value, rfid_value
            )
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            app.save_log(
                workbook,
                timestamp,
                potentiometer_value,
                cam_value,
                rfid_value,
                0,
                indexExit,
            )
            capture_photo(timestamp)

            ultimo_tempo = tempo_atual
